from .regex_pattern import extract_imps
from .update_sheet import update_google_sheets
from .spacy_normalize import is_same_name, description_contains_category
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.response import Response
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell
from rest_framework.views import APIView
from datetime import datetime, timedelta
from openpyxl import load_workbook
from rest_framework import status
from pathlib import Path
from copy import copy
from collections import Counter
import pandas as pd
import openpyxl
import re
from django.conf import settings
import os


class FormatStatement(APIView):
    parser_classes = [MultiPartParser, FormParser]

    def post(self, request, *args, **kwargs):
        excel_files = request.FILES.getlist("files")

        AMOUNT_TOLERANCE = 100.0
        PROCESSED_DIR = "Matched_Statemants"

        TITLE_WORDS = {"MR", "MRS", "MS", "MISS", "DR", "SHRI", "SMT"}
        FIRM_KEYWORDS = {
            "AGENCY", "AGENCIES", "ENTERPRISE", "ENTERPRISES", "TRADERS", "TRADING",
            "INDUSTRIES", "INDUSTRY", "CO", "COMPANY", "LLP", "LTD", "LIMITED",
            "ASSOCIATES", "TRUST", "FOUNDATION", "CENTRE", "CENTER",
            "STORE", "STORES", "SHOP", "SOLUTIONS", "PVT", "PRIVATE"
        }

        SHORT_BANK_NAMES = {
            "Axis Bank, India": "AXIS",
            "Bank of Baroda, India": "BOB",
            "Bank of India, India": "BOI",
            "Bandhan Bank, India": "BDBL",
            "Canara Bank, India": "CNRB",
            "Catholic Syrian Bank, India": "CSB",
            "Citibank, India": "CITI",
            "City Union Bank, India": "CUB",
            "DBS Bank, India": "DBS",
            "Dhanalakshmi Bank Ltd., India": "DLXB",
            "Deutsche Bank, India": "DEUT",
            "Equitas Bank, India": "ESFB",
            "Federal Bank, India": "FDRL",
            "HDFC Bank, India": "HDFC",
            "ICICI Bank, India": "ICICI",
            "IDFC First Bank, India": "IDFC",
            "IDFC FIRST Bank, India": "IDFC",
            "Indian Bank, India": "IDIB",
            "IDBI, India": "IDBI",
            "Indian Overseas Bank, India": "IOB",
            "IndusInd Bank, India": "INDB",
            "Jana Small Finance Bank Ltd, India": "JSFB",
            "Karur Vysya Bank, India": "KVB",
            "Karnataka Bank, India": "KARB",
            "Kotak Mahindra Bank, India": "KKBK",
            "Punjab National Bank, India": "PNB",
            "RBL Bank, India": "RBL",
            "RBL (Ratnakar) Bank, India": "RBL",
            "South Indian Bank, India": "SIB",
            "State Bank of India, India": "SBI",
            "Standard Chartered Bank, India": "",
            "Tamilnad Mercantile Bank Ltd, India": "TMB",
            "Tamilnad Mercantile Bank Ltd., India": "TMB",
            "Union Bank of India, India": "UBI",
            "Ujjivan Bank, India": "UJVN",
            "UCO Bank, India": "UCO",
            "Yes Bank, India": "YBL",
        }

        # -------------------- Helpers -------------------- #
        def get_downloads():
            home = Path.home()
            downloads = home / "Downloads"
            return downloads

        def get_sheet_name(df):
            wanted = [
                "Name of the Account Holder",
                "Name of the Bank",
                "Account Number",
                "Account Type",
            ]
            sub = df[df[1].isin(wanted)]
            info = dict(zip(sub[1], sub[2]))
            account_number = str(info['Account Number']).strip()
            last_4_digits = account_number[-4:] if len(account_number) >= 4 else account_number
            
            sheet_name = (
                f"XNS-{SHORT_BANK_NAMES[info['Name of the Bank']]}-"
                f"{last_4_digits}"
            )
            return sheet_name, info["Name of the Account Holder"]

        def preprocess_category(category):
            return (
                str(category)
                .replace("Transfer from", "")
                .replace("Transfer to", "")
                .strip()
            )

        def get_acc_type(bal):
            non_zero = bal[bal != 0]
            neg_count = (non_zero < 0).sum()
            pos_count = (non_zero > 0).sum()

            if len(non_zero) > 0 and neg_count > pos_count:
                acc_type = "OD"
            else:
                acc_type = "CA"
            return acc_type

        def get_numbers(desc):
            text = str(desc)
            
            # First, try the original pattern for masked accounts
            text_no_space = re.sub(r"\s+", "", text)
            matches = re.findall(r"(?:X|x){4,}\d{3,}", text_no_space)
            
            # If no matches found, check for specific account number patterns
            if not matches:
                pattern1 = r'MOB/SELFFT/[^/]+/(\d{12,16})'
                matches1 = re.findall(pattern1, text, re.IGNORECASE)
                matches.extend(matches1)
                
                pattern2 = r'Self Transfer (?:to|from) (\d{12,16})'
                matches2 = re.findall(pattern2, text, re.IGNORECASE)
                matches.extend(matches2)
            
            return matches

        def extract_acc_suffix_from_key(key):
            digit_groups = re.findall(r"\d+", key)
            if not digit_groups:
                return None
            return digit_groups[-1][-4:]

        def extract_bank_name_from_sheet(sheet_name):
            splitted_sheet = str(sheet_name).split("-")
            filtered_list = [
                item
                for item in splitted_sheet
                if not item.isdigit() and item not in ["XNS", "CA", "OD"]
            ]
            return filtered_list[0] if filtered_list else None

        def normalize_date(date_val):
            if pd.isna(date_val):
                return None
            try:
                if isinstance(date_val, str):
                    date_part = date_val.split()[0] if " " in date_val else date_val
                    return pd.to_datetime(date_part).normalize()
                elif isinstance(date_val, datetime):
                    return date_val.normalize()
                else:
                    return pd.to_datetime(date_val).normalize()
            except Exception:
                return None

        def normalize_name(name: str) -> str:
            """Normalize company names by removing common variations"""
            if not name or not isinstance(name, str):
                return ""
            
            s = str(name).upper().strip()
            
            # Remove common prefixes
            s = re.sub(r"^M/S[.\s]*", "", s)
            s = re.sub(r"^THE\s+", "", s)
            
            # Standardize company suffixes
            s = re.sub(r"\b(PVT|PRIVATE)\s+(LTD|LIMITED)\b", "PVT LTD", s)
            s = re.sub(r"\bPVT\.?\s*LTD\.?\b", "PVT LTD", s)
            s = re.sub(r"\bPRIVATE\s+LIMITED\b", "PVT LTD", s)
            s = re.sub(r"\bLTD\.?\b", "LTD", s)
            s = re.sub(r"\bCO\.?\b", "CO", s)
            s = re.sub(r"\bAND\b", "&", s)
            
            # Remove special characters
            s = re.sub(r"[^\w\s&]", " ", s)
            s = re.sub(r"\s+", " ", s).strip()
            
            # Remove common titles
            parts = [p for p in s.split() if p not in TITLE_WORDS]
            return " ".join(parts)

        def get_party_type(raw_name: str) -> str:
            norm = normalize_name(raw_name)
            if not norm:
                return "OTHER"
            
            tokens = norm.split()
            if any(tok in FIRM_KEYWORDS for tok in tokens):
                return "COMPANY"
            if len(tokens) >= 4:
                return "COMPANY"
            if 1 <= len(tokens) <= 3 and all(t.isalpha() for t in tokens):
                return "PERSON"
            return "OTHER"

        def extract_core_name(name: str) -> str:
            """Extract the core business name by removing company suffixes"""
            norm = normalize_name(name)
            if not norm:
                return ""
            
            # Remove common company suffixes
            suffixes = ["PVT LTD", "LTD", "LIMITED", "CO", "LLP", "PVT", "PRIVATE"]
            core = norm
            for suffix in suffixes:
                pattern = r"\s+" + re.escape(suffix) + r"\b"
                core = re.sub(pattern, "", core, flags=re.IGNORECASE)
            
            # Remove any standalone ampersands
            core = re.sub(r"\s*&\s*", " ", core).strip()
            
            return core

        def same_entity(name1: str, name2: str) -> bool:
            """Check if two names refer to the same entity"""
            if not name1 or not name2:
                return False
            
            n1 = normalize_name(name1)
            n2 = normalize_name(name2)
            
            # If normalized names are identical
            if n1 == n2:
                return True
            
            # Extract core names (without company suffixes)
            core1 = extract_core_name(name1)
            core2 = extract_core_name(name2)
            
            # If core names are identical
            if core1 and core2 and core1 == core2:
                return True
            
            # Split into tokens
            t1 = set(core1.split())
            t2 = set(core2.split())
            
            # If both are empty after processing
            if not t1 or not t2:
                return False
            
            # Check for significant overlap
            common = t1.intersection(t2)
            
            # For companies, check if most words match
            if get_party_type(name1) == "COMPANY" and get_party_type(name2) == "COMPANY":
                # Get the minimum length (excluding common filler words)
                filler_words = {"THE", "AND", "OF", "FOR", "WITH"}
                t1_filtered = {t for t in t1 if t not in filler_words}
                t2_filtered = {t for t in t2 if t not in filler_words}
                common_filtered = {t for t in common if t not in filler_words}
                
                min_len = min(len(t1_filtered), len(t2_filtered))
                if min_len == 0:
                    return False
                
                # If most words match
                match_ratio = len(common_filtered) / min_len
                if match_ratio >= 0.7:  # 70% match
                    return True
                
                # Check for acronym/abbreviation match
                # E.g., "GTIMES" vs "G TIMES"
                core1_no_space = core1.replace(" ", "")
                core2_no_space = core2.replace(" ", "")
                if core1_no_space == core2_no_space:
                    return True
            
            return False

        def infer_transfer_type(name_from: str, name_to: str) -> str:
            """Infer the type of transfer between two entities"""
            t1 = get_party_type(name_from)
            t2 = get_party_type(name_to)

            if t1 == t2 and t1 in {"COMPANY", "PERSON"}:
                if same_entity(name_from, name_to):
                    return "INB TRF"
                else:
                    return "SIS CON"

            if {"COMPANY", "PERSON"} == {t1, t2}:
                return "SIS CON"

            return "OTHERS"

        # Add this helper function
        def is_valid_xns_sheet_name(sheet_name):
            """
            Check if a sheet name is a valid XNS sheet format
            Valid formats:
            - XNS-BANKCODE-ACCNUM-PRODUCT (e.g., XNS-IOB-4241-OD)
            - XNS-ACCNUM-BANKCODE-PRODUCT (e.g., XNS-4241-IOB-OD)
            - BANKCODE-ACCNUM-PRODUCT-XNS (e.g., IOB-4241-OD-XNS)
            - BANKCODE-ACCNUM-PRODUCT (e.g., IOB-4241-OD)
            """
            sheet_name = str(sheet_name).upper().replace(' ', '')
            
            # Skip obviously invalid names
            if sheet_name == "XNS":
                return False
            
            # Check for valid patterns
            valid_patterns = [
                r'^XNS[-_][A-Z]{3,5}[-_]\d{3,4}[-_][A-Z]{2}$',  # XNS-BANKCODE-ACCNUM-PRODUCT
                r'^XNS[-_]\d{3,4}[-_][A-Z]{3,5}[-_][A-Z]{2}$',  # XNS-ACCNUM-BANKCODE-PRODUCT
                r'^[A-Z]{3,5}[-_]\d{3,4}[-_][A-Z]{2}[-_]XNS$',  # BANKCODE-ACCNUM-PRODUCT-XNS
                r'^[A-Z]{3,5}[-_]\d{3,4}[-_][A-Z]{2}$',         # BANKCODE-ACCNUM-PRODUCT
                r'^\d{3,4}[-_][A-Z]{3,5}[-_][A-Z]{2}$',         # ACCNUM-BANKCODE-PRODUCT
            ]
            
            for pattern in valid_patterns:
                if re.match(pattern, sheet_name):
                    return True
            
            return False

        def canonical_sheet_id(name: str) -> str:
            """
            Normalise XNS sheet names so that:
            - 'XNS-SBI-0987-CA' and 'SBI-0987-CA-XNS' both become 'SBI-0987-CA'
            - Also handles BOB-361-CA, 361-BOB-CA, etc.
            - Returns empty string for invalid sheet names
            """
            name = str(name).strip().upper().replace(' ', '')
            
            # If the name is just "XNS" or doesn't look valid, return empty string
            if name == "XNS" or re.search(r'^XNS[-_]\d+$', name):
                print(f"❌ Invalid sheet name in canonical_sheet_id: '{name}'")
                return ""
            
            # Debug: show what we're processing
            print(f"🔍 canonical_sheet_id processing: '{name}'")
            
            # Pattern 1: XNS-BANKCODE-ACCNUM-PRODUCT (e.g., XNS-SBI-0987-CA)
            pattern1 = r'^XNS[-_]?([A-Z]{3,5})[-_]?(\d{3,4})[-_]?([A-Z]{2})$'
            match1 = re.search(pattern1, name)
            if match1:
                bank_code, acc_num, product = match1.groups()
                # Add X prefix for 3-digit accounts
                if len(acc_num) == 3:
                    acc_num = 'X' + acc_num
                result = f"{bank_code}-{acc_num}-{product}"
                print(f"   Pattern 1 matched: {result}")
                return result
            
            # Pattern 2: XNS-ACCNUM-BANKCODE-PRODUCT (e.g., XNS-0987-SBI-CA)
            pattern2 = r'^XNS[-_]?(\d{3,4})[-_]?([A-Z]{3,5})[-_]?([A-Z]{2})$'
            match2 = re.search(pattern2, name)
            if match2:
                acc_num, bank_code, product = match2.groups()
                # Add X prefix for 3-digit accounts
                if len(acc_num) == 3:
                    acc_num = 'X' + acc_num
                result = f"{bank_code}-{acc_num}-{product}"
                print(f"   Pattern 2 matched: {result}")
                return result
            
            # Pattern 3: BANKCODE-ACCNUM-PRODUCT-XNS (e.g., SBI-0987-CA-XNS)
            pattern3 = r'^([A-Z]{3,5})[-_]?(\d{3,4})[-_]?([A-Z]{2})[-_]?XNS$'
            match3 = re.search(pattern3, name)
            if match3:
                bank_code, acc_num, product = match3.groups()
                # Add X prefix for 3-digit accounts
                if len(acc_num) == 3:
                    acc_num = 'X' + acc_num
                result = f"{bank_code}-{acc_num}-{product}"
                print(f"   Pattern 3 matched: {result}")
                return result
            
            # Pattern 4: BANKCODE-ACCNUM-PRODUCT (e.g., SBI-0987-CA, BOB-361-CA)
            pattern4 = r'^([A-Z]{3,5})[-_]?(\d{3,4})[-_]?([A-Z]{2})$'
            match4 = re.search(pattern4, name)
            if match4:
                bank_code, acc_num, product = match4.groups()
                # Add X prefix for 3-digit accounts
                if len(acc_num) == 3:
                    acc_num = 'X' + acc_num
                result = f"{bank_code}-{acc_num}-{product}"
                print(f"   Pattern 4 matched: {result}")
                return result
            
            # Pattern 5: ACCNUM-BANKCODE-PRODUCT (e.g., 0987-SBI-CA)
            pattern5 = r'^(\d{3,4})[-_]?([A-Z]{3,5})[-_]?([A-Z]{2})$'
            match5 = re.search(pattern5, name)
            if match5:
                acc_num, bank_code, product = match5.groups()
                # Add X prefix for 3-digit accounts
                if len(acc_num) == 3:
                    acc_num = 'X' + acc_num
                result = f"{bank_code}-{acc_num}-{product}"
                print(f"   Pattern 5 matched: {result}")
                return result
            
            # If none of the patterns match, return empty string
            print(f"   No pattern matched, returning empty string for: '{name}'")
            return ""

        # -------------------- Generate Summary Report -------------------- #
        def generate_summary_report(matched_df, df_storage, separate_canon_map, final_canon_map, final_file_label, acc_name_storage):
            summary_data = []
            
            print("🔍 Generating summary report...")
            print(f"Files in matched_df: {list(matched_df.keys())}")
            print(f"Files in separate_canon_map: {list(separate_canon_map.keys())}")
            
            for canon, sep_sheet in separate_canon_map.items():
                final_sheet = final_canon_map.get(canon)
                if final_sheet is None:
                    print(f"⚠️ Skipping {sep_sheet} - no matching final sheet")
                    continue

                # Ensure the file exists in matched_df
                if sep_sheet not in matched_df:
                    print(f"⚠️ {sep_sheet} not in matched_df, skipping")
                    continue

                auto_df_full = matched_df.get(sep_sheet)
                manual_df_full = df_storage.get(final_sheet)

                if auto_df_full is None or manual_df_full is None:
                    print(f"⚠️ Skipping {sep_sheet} - missing data (auto: {auto_df_full is not None}, manual: {manual_df_full is not None})")
                    continue

                # Get counts
                total_entries_manual = len(manual_df_full)
                total_entries_software = len(auto_df_full)
                
                # Count INB TRF/SIS CON matches in both files
                manual_inb_sis = manual_df_full[
                    manual_df_full["TYPE"].astype(str).str.upper().isin(["INB TRF", "SIS CON"])
                ]
                software_inb_trf = auto_df_full[
                    auto_df_full["TYPE"].astype(str).str.upper().isin(["INB TRF", "SIS CON"])
                ]
                
                manual_matched = len(manual_inb_sis)
                software_matched = len(software_inb_trf)
                
                # Calculate percentage
                if manual_matched > 0:
                    percentage = (software_matched / manual_matched) * 100
                else:
                    percentage = 0
                
                # Generate file name in format: "safe_acc_name-safe_name"
                acc_name = acc_name_storage.get(sep_sheet, "")
                safe_acc_name = acc_name.replace("/", "_")
                
                # Extract safe_name from sheet name (remove XNS prefix)
                safe_name_parts = sep_sheet.split("-")
                if "XNS" in safe_name_parts:
                    safe_name_parts.remove("XNS")
                safe_name = "-".join(safe_name_parts)
                
                file_name = f"{safe_acc_name}-{safe_name}"
                
                # Extract bank name from sheet name
                bank_value = extract_bank_name_from_sheet(sep_sheet)
                bank_name = [key for key, value in SHORT_BANK_NAMES.items() if value == bank_value][0]
                
                summary_data.append({
                    "File Name": file_name,
                    "Bank Name": bank_name,
                    "Total Entries (Manual)": total_entries_manual,
                    "Total Entries (Software)": total_entries_software,
                    "Manual Matched": manual_matched,
                    "Software Matched": software_matched,
                    "Percentage": f"{percentage:.2f}%"
                })
                
                print(f"✅ Added to summary: {file_name}")
            
            print(f"📊 Summary report generated for {len(summary_data)} files")
            return summary_data

        # -------------------- Basic file checks -------------------- #
        if not excel_files:
            return Response(
                {"error": "No files uploaded. Please upload at least 3 .xlsx files (2 statements + 1 final)."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        separate_files = []
        final_file_obj = None

        for f in excel_files:
            name = str(f.name)
            ext = Path(name).suffix.lower()
            if ext != ".xlsx":
                return Response(
                    {
                        "error": "Only .xlsx files are allowed.",
                        "invalid_file": name,
                    },
                    status=status.HTTP_400_BAD_REQUEST,
                )

            if "final" in name.lower():
                if final_file_obj is not None:
                    return Response(
                        {
                            "error": "Multiple 'final' files detected. Please upload exactly one final workbook (.xlsx)."
                        },
                        status=status.HTTP_400_BAD_REQUEST,
                    )
                final_file_obj = f
            else:
                separate_files.append(f)

        if final_file_obj is None:
            return Response(
                {
                    "error": "Missing final workbook. "
                             "Upload exactly 1 final .xlsx file (name containing 'final') "
                             "and at least 2 separate statement .xlsx files."
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        if len(separate_files) < 2:
            return Response(
                {
                    "error": f"At least 2 separate statement .xlsx files (without 'final' in the name) "
                             f"are required. Only {len(separate_files)} provided."
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        final_upload = final_file_obj

        # -------------------- Load separate (non-final) files -------------------- #
        bank_data_storage = {}
        analysis_storage = {}
        acc_name_storage = {}
        statement_storage = {}

        for file in separate_files:
            analysis_df = pd.read_excel(file, sheet_name="Analysis", header=None)
            statement_df = pd.read_excel(
                file, sheet_name="Statements Considered", header=None
            )
            sheet_name, acc_name = get_sheet_name(analysis_df)

            xns_df = pd.read_excel(file, sheet_name="Xns")
            acc_type = get_acc_type(xns_df["Balance"])
            sheet_name = f"{sheet_name}-{acc_type}"

            xns_df["Amount"] = (
                xns_df["Amount"]
                    .astype(str)
                    .str.replace(",", "", regex=False)
                    .astype(float)
            )
            xns_df["DR"] = xns_df["Amount"].where(xns_df["Type"] == "Debit")
            xns_df["CR"] = xns_df["Amount"].where(xns_df["Type"] == "Credit")
            xns_df["Type"] = ""

            new_df = xns_df.drop(columns=["Amount"]).rename(
                columns={
                    "Sl. No. ": "Sl. No.",
                    "Cheque No.": "Cheque_No",
                    "Type": "TYPE",
                }
            )

            new_df["MONTH"] = new_df["Date"].dt.strftime("%b").str.upper()
            new_df["Date"] = pd.to_datetime(new_df["Date"], errors="coerce")
            new_df["Category"] = new_df["Category"].apply(preprocess_category)
            cols = [
                "Sl. No.",
                "Date",
                "MONTH",
                "TYPE",
                "Cheque_No",
                "Category",
                "Description",
                "DR",
                "CR",
                "Balance",
            ]
            new_df = new_df[cols]

            bank_data_storage[sheet_name] = new_df
            acc_name_storage[sheet_name] = acc_name

            wb_src = load_workbook(file, data_only=False)
            analysis_storage[sheet_name] = wb_src["Analysis"]
            statement_storage[sheet_name] = wb_src["Statements Considered"]

        print(f"📁 Loaded {len(bank_data_storage)} separate files: {list(bank_data_storage.keys())}")

        # -------------------- Validate final XNS sheets vs separate files -------------------- #
        all_sheets = pd.ExcelFile(final_upload)
        final_xns_sheets = [s for s in all_sheets.sheet_names if "XNS" in s.upper()]

        if not final_xns_sheets:
            return Response(
                {
                    "error": "Final workbook does not contain any XNS sheets. "
                            "Please ensure the final file has XNS sheets matching the separate files."
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        def reformat_final_sheet_name(sheet_name):
            # Remove all spaces and convert to uppercase
            sheet_name = str(sheet_name).replace(" ", "").strip().upper()
            
            print(f"🔍 Processing sheet name: '{sheet_name}'")
            
            # Pattern 1: BANKCODE-ACCNUM-PRODUCT (your final sheet format - needs X prefix)
            # This matches BOB-361-CA, HDFC-123-CA, etc.
            pattern1 = r'^([A-Z]{3,5})[-_]?(\d{3})[-_]?([A-Z]{2})$'
            match1 = re.search(pattern1, sheet_name)
            
            if match1:
                bank_code, acc_num, product = match1.groups()
                
                # Add 'X' prefix to 3-digit account number
                acc_num = 'X' + acc_num
                
                # Reformat to: BANKCODE-XACCNUM-PRODUCT
                new_name = f"{bank_code}-{acc_num}-{product}"
                print(f"🔁 Reformatted final sheet (pattern 1): '{sheet_name}' -> '{new_name}'")
                return new_name
            
            # Pattern 2: BANKCODE-ACCNUM-PRODUCT (4-digit account - no X prefix needed)
            pattern2 = r'^([A-Z]{3,5})[-_]?(\d{4})[-_]?([A-Z]{2})$'
            match2 = re.search(pattern2, sheet_name)
            
            if match2:
                bank_code, acc_num, product = match2.groups()
                # 4-digit account, no X prefix needed
                new_name = f"{bank_code}-{acc_num}-{product}"
                print(f"🔁 Reformatted final sheet (pattern 2): '{sheet_name}' -> '{new_name}'")
                return new_name
            
            # Pattern 3: XNS-ACCNUM-BANKCODE-PRODUCT (original format) - FIXED THIS PATTERN
            pattern3 = r'^XNS[-_]?(\d{3,4})[-_]?([A-Z]{3,5})[-_]?([A-Z]{2})$'
            match3 = re.search(pattern3, sheet_name)
            
            if match3:
                acc_num, bank_code, product = match3.groups()
                
                # Add 'X' prefix if account number is 3 digits
                if len(acc_num) == 3:
                    acc_num = 'X' + acc_num
                
                # Reformat to: XNS-BANKCODE-ACCNUM-PRODUCT
                new_name = f"XNS-{bank_code}-{acc_num}-{product}"
                print(f"🔁 Reformatted final sheet (pattern 3): '{sheet_name}' -> '{new_name}'")
                return new_name
            
            # Pattern 4: BANKCODE-ACCNUM-PRODUCT-XNS (your new format)
            pattern4 = r'^([A-Z]{3,5})[-_]?(\d{3,4})[-_]?([A-Z]{2})[-_]?XNS$'
            match4 = re.search(pattern4, sheet_name)
            
            if match4:
                bank_code, acc_num, product = match4.groups()
                
                # Add 'X' prefix if account number is 3 digits
                if len(acc_num) == 3:
                    acc_num = 'X' + acc_num
                
                # Reformat to: XNS-BANKCODE-ACCNUM-PRODUCT
                new_name = f"XNS-{bank_code}-{acc_num}-{product}"
                print(f"🔁 Reformatted final sheet (pattern 4): '{sheet_name}' -> '{new_name}'")
                return new_name
            
            # Pattern 5: XNS-BANKCODE-ACCNUM-PRODUCT (already correct format)
            pattern5 = r'^XNS[-_]?([A-Z]{3,5})[-_]?(\d{3,4})[-_]?([A-Z]{2})$'
            match5 = re.search(pattern5, sheet_name)
            
            if match5:
                bank_code, acc_num, product = match5.groups()
                
                # Add 'X' prefix if account number is 3 digits
                if len(acc_num) == 3:
                    acc_num = 'X' + acc_num
                
                # Already in correct format, just ensure X prefix
                new_name = f"XNS-{bank_code}-{acc_num}-{product}"
                print(f"🔁 Reformatted final sheet (pattern 5): '{sheet_name}' -> '{new_name}'")
                return new_name
            
            # If it's already in the correct format or doesn't match, return as is
            print(f"ℹ️  No reformatting needed for: '{sheet_name}'")
            return sheet_name

        # Reformat final sheet names to match processed file format
        print("=== REFORMATTING FINAL SHEET NAMES ===")
        reformatted_final_sheets = []
        df_storage = {}
        
        for sheet in all_sheets.sheet_names:
            # Only process valid XNS sheets
            if "XNS" in sheet.upper() and is_valid_xns_sheet_name(sheet):
                # Reformat the sheet name to match processed file format
                reformatted_sheet = reformat_final_sheet_name(sheet)
                reformatted_final_sheets.append(reformatted_sheet)
                
                # Load the data with original sheet name, but store with reformatted name
                df = pd.read_excel(final_upload, sheet_name=sheet)
                for col in list(df.columns):
                    if "Unnamed" in str(col):
                        df.drop(columns=[col], inplace=True)
                df_storage[reformatted_sheet] = df
                print(f"✅ Loaded valid XNS sheet: '{sheet}' -> stored as '{reformatted_sheet}'")
            elif "XNS" in sheet.upper():
                print(f"⚠️  Skipping invalid XNS sheet: '{sheet}' (not a valid format)")
            else:
                print(f"ℹ️  Skipping non-XNS sheet: '{sheet}'")

        # Update final_xns_sheets to only include valid sheets
        final_xns_sheets = reformatted_final_sheets
        print(f"📋 Valid final XNS sheets: {final_xns_sheets}")

        separate_sheet_names = list(bank_data_storage.keys())
        
        # Create canonical maps, filtering out invalid sheets
        separate_canon_map = {}
        for s in separate_sheet_names:
            canon = canonical_sheet_id(s)
            if canon:  # Only add if canonical_sheet_id returned a non-empty string
                separate_canon_map[canon] = s
            else:
                print(f"⚠️  Skipping invalid separate sheet: '{s}'")

        final_canon_map = {}
        for s in final_xns_sheets:
            canon = canonical_sheet_id(s)
            if canon:  # Only add if canonical_sheet_id returned a non-empty string
                final_canon_map[canon] = s
            else:
                print(f"⚠️  Skipping invalid final sheet: '{s}'")

        # Filter out empty strings from canon sets
        separate_canon_set = {name for name in set(separate_canon_map.keys()) if name}
        final_canon_set = {name for name in set(final_canon_map.keys()) if name}

        print("=== CANONICAL MAPPING AFTER REFORMATTING ===")
        print("Separate canon map:", separate_canon_map)
        print("Final canon map:", final_canon_map)
        print("Separate canon set:", separate_canon_set)
        print("Final canon set:", final_canon_set)

        if separate_canon_set != final_canon_set:
            missing_in_separate = sorted(final_canon_set - separate_canon_set)
            missing_in_final = sorted(separate_canon_set - final_canon_set)
            
            # Create clear error message - only show unmatched
            error_message = "❌ SHEET MATCHING FAILED: "
            
            if missing_in_separate:
                for sheet in missing_in_separate:
                    error_message += f"{sheet}"
                    error_message += ", "
            
            if missing_in_final:
                for sheet in missing_in_final:
                    error_message += f"{sheet}"
                    error_message += ", "
            
            return Response(
                {
                    "error": error_message,
                    "details": {
                        "software_sheets": sorted(list(separate_canon_set)),  # All software sheets
                        "final_sheets": sorted(list(final_canon_set)),        # All final sheets
                        "unmatched_software_sheets": missing_in_final,        # Sheets only in software
                        "unmatched_final_sheets": missing_in_separate,        # Sheets only in final
                    }
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        # ---------------------- MATCHING LOGIC (between accounts) ---------------------- #
        def compare_files(bank_data_storage):
            # ... [Keep all the existing matching logic code exactly as it is] ...
            # This function remains unchanged from your original code
            all_files = list(bank_data_storage.keys())
            if len(all_files) < 2:
                print(
                    f"I want min 2 files, you uploaded less than 2 files: {all_files}"
                )
                return bank_data_storage

            print(f"Files to process: {all_files}")

            total_matches = 0
            inb_trf_count = 0

            def preprocess_df(df, bank_name):
                if "norm_date" not in df.columns:
                    df["norm_date"] = df["Date"].apply(normalize_date)

                if "CR_val" not in df.columns:
                    cr = pd.to_numeric(df["CR"], errors="coerce").fillna(0.0)
                    df["CR_val"] = cr.abs()

                if "DR_val" not in df.columns:
                    dr = pd.to_numeric(df["DR"], errors="coerce").fillna(0.0)
                    df["DR_val"] = dr.abs()

                if "IMPS" not in df.columns:
                    df["IMPS"] = df["Description"].astype(str).apply(
                        lambda d: extract_imps(d, bank_name)
                    )

                if "NUMBERS" not in df.columns:
                    df["NUMBERS"] = df["Description"].astype(str).apply(get_numbers)

                if "IMPS" in df.columns and "NUMBERS" in df.columns:
                    imps_pos = df.columns.get_loc("IMPS")
                    nums_series = df["NUMBERS"]
                    df.drop(columns=["NUMBERS"], inplace=True)
                    df.insert(imps_pos + 1, "NUMBERS", nums_series)

            def build_lookup_by_date(df, value_col):
                sub = df[df[value_col] > 0].copy()
                sub["key"] = sub["norm_date"]

                lookup = {}
                for idx, key in zip(sub.index, sub["key"]):
                    lookup.setdefault(key, []).append(idx)
                return lookup

            def block_fallback_by_imps(row1, row2):
                imps1 = str(row1.get("IMPS") or "").strip()
                imps2 = str(row2.get("IMPS") or "").strip()

                if not imps1 or not imps2:
                    return False

                desc1 = str(row1.get("Description") or "")
                desc2 = str(row2.get("Description") or "")

                tail1 = imps1[-5:] if len(imps1) > 5 else imps1
                tail2 = imps2[-5:] if len(imps2) > 5 else imps2

                if (tail1 and tail1 in desc2) or (tail2 and tail2 in desc1):
                    return False

                return True

            def find_imps_candidate(row1, df2, lookup_df2, amount_col1, amount_col2):
                row1_imps = row1.get("IMPS")
                desc1 = str(row1.get("Description", ""))
                
                def extract_numbers_5plus_digits(text):
                    """Extract all numbers with 5 or more digits from text"""
                    if not text:
                        return []
                    # Find all sequences of 5 or more digits
                    return re.findall(r'\b\d{5,}\b', text)

                try:
                    amt1 = float(row1[amount_col1])
                except (TypeError, ValueError):
                    return None

                date_key = row1["norm_date"]

                candidate_idx2_list = []
                for offset in (-1, 0, 1):
                    dkey = date_key + timedelta(days=offset)
                    candidate_idx2_list.extend(lookup_df2.get(dkey, []))
                candidate_idx2_list = list(dict.fromkeys(candidate_idx2_list))

                if not candidate_idx2_list:
                    return None

                matches = []
                
                # Extract numbers from desc1 once
                desc1_numbers = extract_numbers_5plus_digits(desc1)
                
                for idx2 in candidate_idx2_list:
                    row2 = df2.loc[idx2]
                    row2_imps = row2.get("IMPS")
                    desc2 = str(row2.get("Description", ""))

                    has_imps1 = bool(row1_imps) and isinstance(row1_imps, str) and len(row1_imps) >= 5
                    has_imps2 = bool(row2_imps) and isinstance(row2_imps, str) and len(row2_imps) >= 5

                    if not (has_imps1 or has_imps2):
                        continue

                    imps_ok = False
                    match_details = ""
                    
                    # CHECK 1: Direct IMPS last 5 digits match
                    if has_imps1 and has_imps2:
                        row1_imps_tail = row1_imps[-5:]  # Get last 5 digits
                        row2_imps_tail = row2_imps[-5:]  # Get last 5 digits
                        
                        if row1_imps_tail == row2_imps_tail:
                            imps_ok = True
                            match_details = f"Direct IMPS match: row1_imps tail={row1_imps_tail} == row2_imps tail={row2_imps_tail}"
                            print(f"✅ {match_details}")

                    # CHECK 2: row1_imps last 5 digits match ending of any number in desc2
                    if (not imps_ok) and has_imps1:
                        desc2_numbers = extract_numbers_5plus_digits(desc2)
                        row1_imps_tail = row1_imps[-5:]  # Get last 5 digits
                        
                        for number in desc2_numbers:
                            if number.endswith(row1_imps_tail):
                                imps_ok = True
                                match_details = f"row1_imps tail={row1_imps_tail} ends desc2 number={number}"
                                print(f"✅ {match_details}")
                                break
                    
                    # CHECK 3: row2_imps last 5 digits match ending of any number in desc1
                    if (not imps_ok) and has_imps2:
                        row2_imps_tail = row2_imps[-5:]  # Get last 5 digits
                        
                        for number in desc1_numbers:
                            if number.endswith(row2_imps_tail):
                                imps_ok = True
                                match_details = f"row2_imps tail={row2_imps_tail} ends desc1 number={number}"
                                print(f"✅ {match_details}")
                                break

                    if not imps_ok:
                        continue

                    try:
                        amt2 = float(row2[amount_col2])
                    except (TypeError, ValueError):
                        continue

                    if abs(amt1 - amt2) <= AMOUNT_TOLERANCE:
                        matches.append(idx2)

                if len(matches) == 1:
                    return matches[0]
                return None

            # def find_self_candidate(row1, df2, lookup_df2, amount_col1, amount_col2, this_acc_name, other_acc_name):
            #     try:
            #         amt1 = float(row1[amount_col1])
            #     except (TypeError, ValueError):
            #         return None

            #     date_key = row1["norm_date"]
            #     acc_candidates = lookup_df2.get(date_key, [])
            #     if not acc_candidates:
            #         return None

            #     # Normalize account names
            #     # this_norm = normalize_name(this_acc_name) if this_acc_name else ""
            #     # other_norm = normalize_name(other_acc_name) if other_acc_name else ""

            #     # Get and normalize category from row1
            #     cat1_raw = str(row1.get("Category", "")).strip()
            #     # cat1_norm = normalize_name(cat1_raw)
                
            #     # Get and normalize description from row1
            #     desc1_raw = str(row1.get("Description", "")).strip()
            #     # desc1_norm = normalize_name(desc1_raw)

            #     matches = []
            #     for idx2 in acc_candidates:
            #         row2 = df2.loc[idx2]
            #         try:
            #             amt2 = float(row2[amount_col2])
            #         except (TypeError, ValueError):
            #             continue
            #         if amt2 != amt1:
            #             continue
                    
            #         cat2_raw = str(row2.get("Category", "")).strip()
            #         desc2_raw = str(row2.get("Description", "")).strip()
                    
            #         is_match = False

            #         if cat1_raw.upper() == 'SELF':
            #             # if cat2_raw.upper() == 'SELF':
            #             #     is_match = True
            #             if is_same_name(this_acc_name, cat2_raw)['same']:
            #                 is_match = True
            #             elif description_contains_category(this_acc_name, desc2_raw)['contains']:
            #                 is_match = True
                        
            #         elif is_same_name(other_acc_name, cat1_raw)['same']:
            #             if cat2_raw.upper() == 'SELF':
            #                 is_match = True
            #             elif is_same_name(this_acc_name, cat2_raw)['same']:
            #                 is_match = True
            #             elif description_contains_category(this_acc_name, desc2_raw)['contains']:
            #                 is_match = True
                        
            #         # elif description_contains_category(other_acc_name, desc1_raw)['contains']:
            #         #     if cat2_raw.upper() == 'SELF':
            #         #         is_match = True
            #         #     elif is_same_name(this_acc_name, cat2_raw)['same']:
            #         #         is_match = True
            #         #     elif description_contains_category(this_acc_name, desc2_raw)['contains']:
            #         #         is_match = True
                    
            #         if is_match:
            #             matches.append(idx2)

            #     if len(matches) == 1:
            #         return matches[0]
            #     return None
            

            def find_self_candidate(row1, df2, lookup_df2, amount_col1, amount_col2, this_acc_name, other_acc_name, df1_key, df2_key):
                try:
                    amt1 = float(row1[amount_col1])
                except (TypeError, ValueError):
                    return None

                date_key = row1["norm_date"]
                acc_candidates = lookup_df2.get(date_key, [])
                if not acc_candidates:
                    return None

                # Get and normalize category from row1
                cat1_raw = str(row1.get("Category", "")).strip()
                
                # Get and normalize description from row1
                desc1_raw = str(row1.get("Description", "")).strip()
                
                # Helper function to extract last 4 digits of account number from sheet key
                def extract_acc_suffix_from_key(key):
                    """Extract last 4 digits of account number from key like XNS-BANKCODE-ACCNUM-PRODUCT"""
                    key_parts = key.split("-")
                    for part in key_parts:
                        # Handle 3-digit accounts with X prefix, 4-digit accounts
                        clean_part = part.replace('X', '')  # Remove X prefix if present
                        if len(clean_part) in [3, 4] and clean_part.isdigit():
                            return clean_part[-4:]  # Get last 4 digits (handles 3-digit by padding)
                    return None
                
                # Get account suffixes for both comparing accounts
                df1_acc_suffix = extract_acc_suffix_from_key(df1_key)  # e.g., "3922" from XNS-HDFC-3922-CA
                df2_acc_suffix = extract_acc_suffix_from_key(df2_key)  # e.g., "9079" from XNS-IDIB-9079-OD
                
                # Check NUMBERS column in row1
                row1_acc_suffixes = []
                row1_numbers = row1.get("NUMBERS", [])
                
                # Only use NUMBERS column if it exists and has data
                if isinstance(row1_numbers, list) and len(row1_numbers) > 0:
                    for num in row1_numbers:
                        if isinstance(num, str) and len(num) >= 4:
                            # Get last 4 digits
                            suffix = num[-4:]
                            if suffix.isdigit():
                                row1_acc_suffixes.append(suffix)
                
                # LOGIC: If account number is present in NUMBERS column, it MUST match one of the comparing accounts
                if row1_acc_suffixes:
                    # Check if any account suffix matches df1 or df2 account
                    account_match_found = False
                    for suffix in row1_acc_suffixes:
                        if (df1_acc_suffix and suffix == df1_acc_suffix) or \
                        (df2_acc_suffix and suffix == df2_acc_suffix):
                            account_match_found = True
                            print(f"✅ Row1 account {suffix} matches comparing account ({df1_acc_suffix} or {df2_acc_suffix})")
                            break
                    
                    if not account_match_found:
                        return None  # VETO: Account present but doesn't match
                
                matches = []
                for idx2 in acc_candidates:
                    row2 = df2.loc[idx2]
                    try:
                        amt2 = float(row2[amount_col2])
                    except (TypeError, ValueError):
                        continue
                    if amt2 != amt1:
                        continue
                    
                    cat2_raw = str(row2.get("Category", "")).strip()
                    desc2_raw = str(row2.get("Description", "")).strip()
                    
                    # Check NUMBERS column in row2
                    row2_acc_suffixes = []
                    row2_numbers = row2.get("NUMBERS", [])
                    
                    # Only use NUMBERS column if it exists and has data
                    if isinstance(row2_numbers, list) and len(row2_numbers) > 0:
                        for num in row2_numbers:
                            if isinstance(num, str) and len(num) >= 4:
                                # Get last 4 digits
                                suffix = num[-4:]
                                if suffix.isdigit():
                                    row2_acc_suffixes.append(suffix)
                    
                    # LOGIC: If account number is present in row2 NUMBERS column, it MUST match one of the comparing accounts
                    if row2_acc_suffixes:
                        account_match_found = False
                        for suffix in row2_acc_suffixes:
                            if (df1_acc_suffix and suffix == df1_acc_suffix) or \
                            (df2_acc_suffix and suffix == df2_acc_suffix):
                                account_match_found = True
                                break
                        
                        if not account_match_found:
                            continue
                    
                    is_match = False

                    if cat1_raw.upper() == 'SELF':
                        if is_same_name(this_acc_name, cat2_raw)['same']:
                            is_match = True
                        elif description_contains_category(this_acc_name, desc2_raw)['contains']:
                            is_match = True
                        
                    elif is_same_name(other_acc_name, cat1_raw)['same']:
                        if cat2_raw.upper() == 'SELF':
                            is_match = True
                        elif is_same_name(this_acc_name, cat2_raw)['same']:
                            is_match = True
                        elif description_contains_category(this_acc_name, desc2_raw)['contains']:
                            is_match = True
                    
                    if is_match:
                        matches.append(idx2)

                if len(matches) == 1:
                    return matches[0]
                return None
            
            def find_etxn_match(row1, df2, lookup_df2, amount_col1, amount_col2, matched_indices_df2):
                """Find eTXN matches based on description pattern, amount, and date."""
                try:
                    amt1 = float(row1[amount_col1])
                except (TypeError, ValueError):
                    return None
                
                # Check if row1 has eTXN pattern and Transfer category
                desc1 = str(row1.get("Description", "")).strip()
                cat1 = str(row1.get("Category", "")).strip().upper()
                
                # Check for eTXN pattern and Transfer category
                has_etxn_pattern1 = re.search(r"(?i)^eTXN\/(?:By|To):(\d+)(?:\/Trf)?", desc1, re.IGNORECASE) is not None
                is_transfer_category1 = cat1 in ["TRANSFER IN", "TRANSFER OUT"]
                
                if not (has_etxn_pattern1 and is_transfer_category1):
                    return None
                
                date_key = row1["norm_date"]
                candidates = lookup_df2.get(date_key, [])
                if not candidates:
                    return None
                
                for idx2 in candidates:
                    if idx2 in matched_indices_df2:
                        continue
                        
                    row2 = df2.loc[idx2]
                    
                    try:
                        amt2 = float(row2[amount_col2])
                    except (TypeError, ValueError):
                        continue
                    
                    if amt2 != amt1:
                        continue
                    
                    # Check if row2 also has eTXN pattern and Transfer category
                    desc2 = str(row2.get("Description", "")).strip()
                    cat2 = str(row2.get("Category", "")).strip().upper()
                    
                    has_etxn_pattern2 = re.search(r"(?i)^eTXN\/(?:By|To):(\d+)(?:\/Trf)?", desc2, re.IGNORECASE) is not None
                    is_transfer_category2 = cat2 in ["TRANSFER IN", "TRANSFER OUT"]
                    
                    if has_etxn_pattern2 and is_transfer_category2:
                        # Extract the eTXN ID to verify it's the same transaction
                        match1 = re.search(r"(?i)^eTXN\/(?:By|To):(\d+)(?:\/Trf)?", desc1, re.IGNORECASE)
                        match2 = re.search(r"(?i)^eTXN\/(?:By|To):(\d+)(?:\/Trf)?", desc2, re.IGNORECASE)
                        
                        if match1 and match2:
                            etxn_id1 = match1.group(1)
                            etxn_id2 = match2.group(1)
                            
                            if etxn_id1 == etxn_id2:
                                # Same eTXN ID - definitely a match
                                return idx2
                            else:
                                # Different eTXN IDs but same amount/date - could still be related
                                # Check if amounts are opposite (one credit, one debit)
                                if (cat1 == "TRANSFER IN" and cat2 == "TRANSFER OUT") or \
                                (cat1 == "TRANSFER OUT" and cat2 == "TRANSFER IN"):
                                    return idx2
                        else:
                            # eTXN pattern exists but no ID extracted - match by amount/date/pattern
                            if (cat1 == "TRANSFER IN" and cat2 == "TRANSFER OUT") or \
                            (cat1 == "TRANSFER OUT" and cat2 == "TRANSFER IN"):
                                return idx2
                
                return None

            def find_acc_candidate(row1, df2, lookup_df2, acc_suffix_df1, acc_suffix_df2, amount_col1, amount_col2):
                date_key = row1["norm_date"]
                acc_candidates = lookup_df2.get(date_key, [])
                if not acc_candidates:
                    return None

                desc1 = str(row1["Description"])
                nums_in_desc1 = get_numbers(desc1)

                try:
                    amt1 = float(row1[amount_col1])
                except (TypeError, ValueError):
                    return None

                acc_suffix_1 = (
                    str(acc_suffix_df1) if acc_suffix_df1 is not None else ""
                )
                acc_suffix_2 = (
                    str(acc_suffix_df2) if acc_suffix_df2 is not None else ""
                )

                exact_candidates = []
                for idx2 in acc_candidates:
                    row2 = df2.loc[idx2]

                    if block_fallback_by_imps(row1, row2):
                        continue

                    try:
                        amt2 = float(row2[amount_col2])
                    except (TypeError, ValueError):
                        continue
                    if amt2 != amt1:
                        continue

                    exact_candidates.append(idx2)

                if not exact_candidates:
                    return None

                matches = []
                for idx2 in exact_candidates:
                    row2 = df2.loc[idx2]
                    desc2 = str(row2["Description"])
                    nums_in_desc2 = get_numbers(desc2)

                    acc_match = False

                    if acc_suffix_2 and nums_in_desc1:
                        if any(num.endswith(acc_suffix_2) for num in nums_in_desc1):
                            acc_match = True

                    if (not acc_match) and acc_suffix_1 and nums_in_desc2:
                        if any(num.endswith(acc_suffix_1) for num in nums_in_desc2):
                            acc_match = True

                    if acc_match:
                        matches.append(idx2)

                if len(matches) == 1:
                    return matches[0]
                return None

            def choose_candidate_for_row(row1, df2, lookup_df2, acc_suffix_df1, acc_suffix_df2, amount_col1, amount_col2, 
                                        other_acc_name, used2, name1, name2):
                idx2 = find_imps_candidate(
                    row1, df2, lookup_df2, amount_col1, amount_col2
                )
                if idx2 is not None and idx2 not in used2:
                    return idx2

                idx2 = find_self_candidate( row1=row1, df2=df2, lookup_df2=lookup_df2, amount_col1=amount_col1, 
                                           amount_col2=amount_col2, this_acc_name=name1, other_acc_name=name2,
                                           df1_key=df1_key, df2_key=df2_key
                )
                if idx2 is not None and idx2 not in used2:
                    return idx2
                
                idx2 = find_etxn_match(
                    row1, df2, lookup_df2, amount_col1, amount_col2, used2
                )
                if idx2 is not None and idx2 not in used2:
                    return idx2


                idx2 = find_acc_candidate(row1, df2, lookup_df2, acc_suffix_df1, acc_suffix_df2, 
                                          amount_col1, amount_col2)
                if idx2 is not None and idx2 not in used2:
                    return idx2

                return None

            for i in range(len(all_files)):
                for j in range(i + 1, len(all_files)):
                    df1_key = all_files[i]
                    df2_key = all_files[j]

                    df1_bank_name = extract_bank_name_from_sheet(df1_key)
                    df2_bank_name = extract_bank_name_from_sheet(df2_key)

                    acc_suffix_df1 = extract_acc_suffix_from_key(df1_key)
                    acc_suffix_df2 = extract_acc_suffix_from_key(df2_key)

                    print(
                        f"\n=== PROCESSING: {df1_key} ({df1_bank_name}) vs "
                        f"{df2_key} ({df2_bank_name}) ==="
                    )

                    df1 = bank_data_storage[df1_key].copy()
                    df2 = bank_data_storage[df2_key].copy()

                    preprocess_df(df1, df1_bank_name)
                    preprocess_df(df2, df2_bank_name)

                    matches_in_pair = 0

                    used_idx1 = set()
                    used_idx2 = set()

                    name1 = acc_name_storage[df1_key]
                    name2 = acc_name_storage[df2_key]

                    def apply_match(idx1, idx2):
                        nonlocal matches_in_pair, total_matches, df1, df2
                        nonlocal inb_trf_count

                        if idx1 in used_idx1 or idx2 in used_idx2:
                            return

                        new_category_df1 = df2_key.split("-")
                        if "XNS" in new_category_df1:
                            new_category_df1.remove("XNS")
                        new_category_df1 = "-".join(new_category_df1)

                        new_category_df2 = df1_key.split("-")
                        if "XNS" in new_category_df2:
                            new_category_df2.remove("XNS")
                        new_category_df2 = "-".join(new_category_df2)

                        type_for_pair = infer_transfer_type(name1, name2)

                        df1.at[idx1, "TYPE"] = type_for_pair
                        df2.at[idx2, "TYPE"] = type_for_pair
                        df1.at[idx1, "Category"] = (
                            f"{name2}-{new_category_df1}"
                        )
                        df2.at[idx2, "Category"] = (
                            f"{name1}-{new_category_df2}"
                        )

                        t = str(type_for_pair).upper()
                        if t in ("INB TRF", "SIS CON"):
                            inb_trf_count += 1

                        used_idx1.add(idx1)
                        used_idx2.add(idx2)
                        matches_in_pair += 1
                        total_matches += 1

                    def process_case(df1_side, df2_lookup, amount_col1, amount_col2, other_acc_name):
                        for idx1, row1 in df1_side.iterrows():
                            if idx1 in used_idx1:
                                continue
                            idx2 = choose_candidate_for_row(row1=row1, df2=df2, lookup_df2=df2_lookup, 
                                                            acc_suffix_df1=acc_suffix_df1, acc_suffix_df2=acc_suffix_df2,
                                                            amount_col1=amount_col1, amount_col2=amount_col2, 
                                                            other_acc_name=other_acc_name, used2=used_idx2,
                                                            name1=name1, name2=name2)
                            if idx2 is not None:
                                apply_match(idx1, idx2)

                    # CASE 1: df1 CR > 0 vs df2 DR > 0
                    lookup_case1 = build_lookup_by_date(df2, "DR_val")
                    df1_case1 = df1[df1["CR_val"] > 0]
                    process_case(df1_case1,lookup_case1,amount_col1="CR_val",amount_col2="DR_val",other_acc_name=name1)

                    # CASE 2: df1 DR > 0 vs df2 CR > 0
                    lookup_case2 = build_lookup_by_date(df2, "CR_val")
                    df1_case2 = df1[df1["DR_val"] > 0]
                    process_case(df1_case2, lookup_case2, amount_col1="DR_val", amount_col2="CR_val", other_acc_name=name1)

                    print(f"📊 Matches in this pair: {matches_in_pair}")

                    for df in (df1, df2):
                        for col in ("norm_date", "CR_val", "DR_val"):
                            if col in df.columns:
                                df.drop(columns=[col], inplace=True)

                    bank_data_storage[df1_key] = df1
                    bank_data_storage[df2_key] = df2

            print("\n=== FINAL SUMMARY ===")
            print(f"TOTAL MATCHES FOUND: {total_matches}")
            print(f"INB TRF COUNT: {inb_trf_count}")

            return bank_data_storage

        # -------------------- Style helpers -------------------- #
        def copy_sheet_with_style(src_ws, wb_out, new_title="Analysis"):
            dest_ws = wb_out.create_sheet(title=new_title)

            for row in src_ws.iter_rows():
                for cell in row:
                    if isinstance(cell, MergedCell):
                        continue

                    dcell = dest_ws.cell(
                        row=cell.row, column=cell.column, value=cell.value
                    )

                    if cell.has_style:
                        dcell.font = copy(cell.font)
                        dcell.fill = copy(cell.fill)
                        dcell.border = copy(cell.border)
                        dcell.number_format = cell.number_format
                        dcell.alignment = copy(cell.alignment)
                        dcell.protection = copy(cell.protection)

            for r_idx, r_dim in src_ws.row_dimensions.items():
                if r_dim.height is not None:
                    dest_ws.row_dimensions[r_idx].height = r_dim.height

            for col_letter, c_dim in src_ws.column_dimensions.items():
                if c_dim.width is not None:
                    dest_ws.column_dimensions[col_letter].width = c_dim.width

            for merged in src_ws.merged_cells.ranges:
                dest_ws.merge_cells(str(merged))

                coord = str(merged)
                top_left_addr = coord.split(":")[0]
                top_left = dest_ws[top_left_addr]
                border = copy(top_left.border)
                fill = copy(top_left.fill)

                for row in dest_ws[coord]:
                    for c in row:
                        c.border = border
                        c.fill = fill

            for col_letter in ["C", "D", "E", "F", "G", "H", "I"]:
                dest_ws.column_dimensions[col_letter].width = 15

            return dest_ws

        def append_sheet_with_style(src_ws, dest_ws, gap_rows=2, col_offset=0):
            dest_last_row = dest_ws.max_row
            row_offset = dest_last_row + gap_rows

            for row in src_ws.iter_rows():
                for cell in row:
                    if isinstance(cell, MergedCell):
                        continue

                    target_row = row_offset + cell.row - 1
                    target_col = cell.column + col_offset

                    dcell = dest_ws.cell(
                        row=target_row, column=target_col, value=cell.value
                    )

                    if cell.has_style:
                        dcell.font = copy(cell.font)
                        dcell.fill = copy(cell.fill)
                        dcell.border = copy(cell.border)
                        dcell.number_format = cell.number_format
                        dcell.alignment = copy(cell.alignment)
                        dcell.protection = copy(cell.protection)

            for r_idx, r_dim in src_ws.row_dimensions.items():
                if r_dim.height is not None:
                    dest_ws.row_dimensions[row_offset + r_idx - 1].height = r_dim.height

            for col_letter, c_dim in src_ws.column_dimensions.items():
                if c_dim.width is None:
                    continue
                src_col_idx = openpyxl.utils.column_index_from_string(col_letter)
                dest_col_idx = src_col_idx + col_offset
                dest_col_letter = get_column_letter(dest_col_idx)
                dest_dim = dest_ws.column_dimensions[dest_col_letter]
                dest_dim.width = max(dest_dim.width or 0, c_dim.width)

            for merged in src_ws.merged_cells.ranges:
                min_row = merged.min_row + row_offset - 1
                max_row = merged.max_row + row_offset - 1
                min_col = merged.min_col + col_offset
                max_col = merged.max_col + col_offset

                coord = (
                    f"{get_column_letter(min_col)}{min_row}:"
                    f"{get_column_letter(max_col)}{max_row}"
                )
                dest_ws.merge_cells(coord)

                top_left_addr = coord.split(":")[0]
                top_left = dest_ws[top_left_addr]
                border = copy(top_left.border)
                fill = copy(top_left.fill)

                for row in dest_ws[coord]:
                    for c in row:
                        c.border = border
                        c.fill = fill
            return dest_ws

        # -------------------- Save processed (automate) files -------------------- #
        automate_files = []

        def save_matched_with_styles(
            dfs_dict, acc_name_storage, highlight_red_positions, highlight_green_positions
        ):
            light_red_fill = PatternFill(
                fill_type="solid", start_color="FFFFCCCC", end_color="FFFFCCCC"
            )
            light_green_fill = PatternFill(
                fill_type="solid", start_color="FFC6EFCE", end_color="FFC6EFCE"
            )

            print(f"💾 Saving {len(dfs_dict)} files with styles...")
            
            for key, df in dfs_dict.items():
                print(f"   Processing: {key}")
                
                src_analysis_ws = analysis_storage[key]
                src_statement_ws = statement_storage[key]

                safe_name_parts = key.split("-")
                if "XNS" in safe_name_parts:
                    safe_name_parts.remove("XNS")
                safe_name = "-".join(safe_name_parts)
                acc_name = acc_name_storage[key]
                safe_acc_name = acc_name.replace("/", "_").replace(",", "")

                download_path = get_downloads()
                base_dir = download_path / PROCESSED_DIR
                base_dir.mkdir(parents=True, exist_ok=True)
                filename = base_dir / f"{safe_acc_name}-{safe_name}.xlsx"

                df_out = df.copy()

                df_out["DR"] = pd.to_numeric(df_out["DR"], errors="coerce").abs()
                df_out["CR"] = pd.to_numeric(df_out["CR"], errors="coerce")

                styler = (
                    df_out.style.set_table_styles(
                        [
                            {
                                "selector": "th",
                                "props": (
                                    "background-color: #002060;"
                                    "color: white;"
                                    "border: 1px solid black;"
                                ),
                            },
                            {"selector": "td", "props": "border: 1px solid black;"},
                        ]
                    )
                    .set_properties(**{"border": "1px solid black"})
                    .set_properties(subset=["Category"], **{"font-weight": "bold"})
                    .format(
                        {
                            "DR": "{:,.2f}",
                            "CR": "{:,.2f}",
                            "Balance": "{:,.2f}",
                        },
                        na_rep="",
                    )
                )

                with pd.ExcelWriter(filename, engine="openpyxl") as writer:
                    sheet_name = key if len(key) <= 31 else "Xns"
                    styler.to_excel(writer, sheet_name=sheet_name, index=False)

                    wb_out = writer.book
                    ws = writer.sheets[sheet_name]

                    header_fill = PatternFill(
                        fill_type="solid",
                        start_color="FF002060",
                        end_color="FF002060",
                    )
                    thin = Side(border_style="thin", color="000000")
                    border = Border(
                        left=thin, right=thin, top=thin, bottom=thin
                    )
                    header_font_arial = Font(
                        name="Arial", size=10, bold=True, color="FFFFFFFF"
                    )
                    body_font_arial = Font(name="Arial", size=10)

                    # header row
                    for cell in ws[1]:
                        cell.fill = header_fill
                        cell.font = header_font_arial
                        cell.border = border
                        cell.alignment = Alignment(
                            horizontal="center", vertical="center"
                        )

                    # body rows
                    for row in ws.iter_rows(
                        min_row=2, max_row=ws.max_row, max_col=ws.max_column
                    ):
                        for cell in row:
                            cell.border = border
                            cell.font = body_font_arial
                            cell.alignment = Alignment(vertical="center")

                    # column map
                    col_index = {c.value: c.column for c in ws[1]}

                    # number formats
                    for r in range(2, ws.max_row + 1):
                        for col_name in ["DR", "CR", "Balance"]:
                            col_idx = col_index.get(col_name)
                            if col_idx is None:
                                continue
                            c = ws.cell(row=r, column=col_idx)
                            c.number_format = "0.00"

                    # balance w/ commas + red negatives
                    bal_col_idx = col_index.get("Balance")
                    if bal_col_idx is not None:
                        for r in range(2, ws.max_row + 1):
                            c = ws.cell(row=r, column=bal_col_idx)
                            if c.value is None:
                                continue
                            try:
                                val = float(str(c.value).replace(",", ""))
                            except ValueError:
                                continue
                            c.number_format = "#,##,##0.00"
                            if val < 0:
                                c.font = c.font.copy(color="FFFF0000")

                    # description wrap
                    desc_col_idx = col_index.get("Description")
                    if desc_col_idx is not None:
                        h = ws.cell(row=1, column=desc_col_idx)
                        h.alignment = Alignment(
                            wrap_text=True,
                            horizontal="center",
                            vertical="center",
                        )
                        for r in range(2, ws.max_row + 1):
                            c = ws.cell(row=r, column=desc_col_idx)
                            c.alignment = Alignment(
                                wrap_text=True, vertical="center"
                            )

                    # date format
                    date_col_idx = col_index.get("Date")
                    if date_col_idx is not None:
                        for r in range(2, ws.max_row + 1):
                            c = ws.cell(row=r, column=date_col_idx)
                            if c.value:
                                c.number_format = "DD-MMM-YY"

                    # widths
                    desired_widths = {
                        "Sl. No.": 8,
                        "Date": 10,
                        "MONTH": 10,
                        "TYPE": 12,
                        "Cheque_No": 12,
                        "Category": 35,
                        "Description": 50,
                        "DR": 15,
                        "CR": 15,
                        "Balance": 18,
                    }
                    for cell in ws[1]:
                        col_name = cell.value
                        col_letter = cell.column_letter
                        w = desired_widths.get(col_name)
                        if w is not None:
                            wb_out[sheet_name].column_dimensions[col_letter].width = w
                    ws.row_dimensions[1].height = 20

                    # 🔴🟢 highlight mismatches in processed file only
                    red_positions = highlight_red_positions.get(key, set())
                    green_positions = highlight_green_positions.get(key, set())
                    n_rows = len(df_out.index)

                    for excel_row in range(2, ws.max_row + 1):
                        pos = excel_row - 2  # df_out position
                        if pos < 0 or pos >= n_rows:
                            continue

                        row_fill = None
                        if pos in green_positions:
                            row_fill = light_green_fill
                        if pos in red_positions:
                            row_fill = light_red_fill  # red overrides if both

                        if row_fill is not None:
                            for cell in ws[excel_row]:
                                cell.fill = row_fill

                    # copy ANALYSIS (first sheet) + Statements into it
                    src_analysis_ws = analysis_storage[key]
                    analysis_ws = copy_sheet_with_style(
                        src_analysis_ws, wb_out, new_title="ANALYSIS"
                    )

                    for col_letter in ["C", "D", "E", "F", "G", "H", "I"]:
                        analysis_ws.column_dimensions[col_letter].width = 15

                    sheets = list(wb_out._sheets)
                    if analysis_ws in sheets:
                        sheets.insert(0, sheets.pop(sheets.index(analysis_ws)))
                        wb_out._sheets = sheets

                    append_sheet_with_style(
                        src_statement_ws, analysis_ws, gap_rows=3, col_offset=1
                    )

                    automate_files.append(filename)
                print(f"✅ Saved styled file: {filename}")

        # --------- run comparison between accounts --------- #
        matched_df = compare_files(bank_data_storage)

        # DEBUG: Check what files we have after comparison
        print("=== AFTER COMPARISON DEBUG ===")
        print(f"Files in matched_df: {list(matched_df.keys())}")
        print(f"Files in bank_data_storage: {list(bank_data_storage.keys())}")
        
        # Ensure ALL files are in matched_df
        all_files_to_process = {}
        for key in bank_data_storage.keys():
            if key in matched_df:
                all_files_to_process[key] = matched_df[key]
                print(f"✅ Using matched data for: {key}")
            else:
                all_files_to_process[key] = bank_data_storage[key]
                print(f"⚠️ Using original data for: {key} (not in matched_df)")

        # ---------- row count summary (separate vs final) ---------- #
        row_count_summary = {}
        final_file_label = Path(final_upload.name).name
        for canon, sep_sheet in separate_canon_map.items():
            final_sheet = final_canon_map.get(canon)
            if final_sheet is None:
                continue
            row_id = f"{final_file_label}::{final_sheet}"
            sep_rows = len(bank_data_storage.get(sep_sheet, pd.DataFrame()))
            fin_rows = len(df_storage.get(final_sheet, pd.DataFrame()))
            row_count_summary[row_id] = {
                "separate_sheet": sep_sheet,
                "final_sheet": final_sheet,
                "separate_rows": int(sep_rows),
                "final_rows": int(fin_rows),
            }

        def compare_inb_sis_rows(auto_df_full, manual_df_full):
            """
            Compare INB TRF/SIS CON rows between processed and final files
            Returns row numbers for mismatches
            """
            # Categories to ignore when comparing OD and CA accounts
            IGNORE_CATEGORIES = {
                "INTEREST", "INTEREST CHARGES", "INTEREST CHARGES REVERSAL",
                "BANK CHARGES", "RETURN", "RETURN CHARGES", "CHARGES",
                "REVERSAL", "PENALTY", "PENALTY CHARGES"
            }
            
            def normalize_category(category):
                """Normalize category for comparison"""
                if pd.isna(category):
                    return ""
                return str(category).strip().upper()
            
            def should_ignore_row(row):
                """Check if a row should be ignored based on category"""
                category = normalize_category(row.get("Category", ""))
                if any(ignore_cat in category for ignore_cat in IGNORE_CATEGORIES):
                    return True
                return False
            
            # Filter processed file for INB TRF/SIS CON only, excluding ignored categories
            auto_inb = auto_df_full[
                auto_df_full["TYPE"].astype(str).str.upper().isin(["INB TRF", "SIS CON"])
            ].copy()
            
            # Filter out rows with ignored categories from processed file
            auto_inb = auto_inb[~auto_inb.apply(should_ignore_row, axis=1)]
            
            # Filter final file for INB TRF and SIS CON, excluding ignored categories
            manual_inb_sis = manual_df_full[
                manual_df_full["TYPE"].astype(str).str.upper().isin(["INB TRF", "SIS CON"])
            ].copy()
            
            # Filter out rows with ignored categories from final file
            manual_inb_sis = manual_inb_sis[~manual_inb_sis.apply(should_ignore_row, axis=1)]
            
            print(f"  After filtering ignored categories:")
            print(f"    Processed rows: {len(auto_inb)} (original: {len(auto_df_full)})")
            print(f"    Final rows: {len(manual_inb_sis)} (original: {len(manual_df_full)})")
            
            # Show which categories were filtered
            auto_filtered = auto_df_full[
                auto_df_full["TYPE"].astype(str).str.upper().isin(["INB TRF", "SIS CON"])
            ].copy()
            auto_filtered = auto_filtered[auto_filtered.apply(should_ignore_row, axis=1)]
            
            if len(auto_filtered) > 0:
                print(f"  Filtered categories from processed file:")
                for category in auto_filtered["Category"].unique():
                    print(f"    - {category}")
            
            manual_filtered = manual_df_full[
                manual_df_full["TYPE"].astype(str).str.upper().isin(["INB TRF", "SIS CON"])
            ].copy()
            manual_filtered = manual_filtered[manual_filtered.apply(should_ignore_row, axis=1)]
            
            if len(manual_filtered) > 0:
                print(f"  Filtered categories from final file:")
                for category in manual_filtered["Category"].unique():
                    print(f"    - {category}")

            def create_comparison_key(row):
                """Create key using Date and Description only"""
                date_val = normalize_date(row["Date"])
                desc = str(row["Description"]).strip().lower()
                return f"{date_val}|{desc}"

            # Create sets of keys for comparison
            auto_keys = set()
            manual_keys = set()

            # Store row numbers for each key
            auto_key_to_rows = {}
            manual_key_to_rows = {}

            # Process auto (processed) file
            for idx, row in auto_inb.iterrows():
                key = create_comparison_key(row)
                auto_keys.add(key)
                if key not in auto_key_to_rows:
                    auto_key_to_rows[key] = []
                # Store actual row number (Excel row number = index + 2)
                auto_key_to_rows[key].append(idx + 2)

            # Process manual (final) file
            for idx, row in manual_inb_sis.iterrows():
                key = create_comparison_key(row)
                manual_keys.add(key)
                if key not in manual_key_to_rows:
                    manual_key_to_rows[key] = []
                # Store actual row number (Excel row number = index + 2)
                manual_key_to_rows[key].append(idx + 2)

            # Find mismatches
            auto_only_keys = auto_keys - manual_keys
            manual_only_keys = manual_keys - auto_keys

            # Get row numbers for mismatches
            auto_only_rows = []
            for key in auto_only_keys:
                auto_only_rows.extend(auto_key_to_rows[key])

            manual_only_rows = []
            for key in manual_only_keys:
                manual_only_rows.extend(manual_key_to_rows[key])

            return {
                "auto_inb_count": len(auto_inb),
                "manual_inb_count": len(manual_inb_sis),
                "auto_only_rows": sorted(auto_only_rows),
                "manual_only_rows": sorted(manual_only_rows),
                "auto_only_count": len(auto_only_rows),
                "manual_only_count": len(manual_only_rows),
                "filtered_categories": list(IGNORE_CATEGORIES),
                "auto_filtered_count": len(auto_filtered),
                "manual_filtered_count": len(manual_filtered),
            }

        mismatch_summary = {}
        highlight_red_positions = {}
        highlight_green_positions = {}
        any_mismatch = False

        print("=== VALIDATING SHEET MATCHING BEFORE COMPARISON ===")

        # Check if all separate files have matching final sheets
        missing_final_sheets = []
        for canon, sep_sheet in separate_canon_map.items():
            final_sheet = final_canon_map.get(canon)
            if final_sheet is None:
                missing_final_sheets.append(sep_sheet)
                print(f"❌ No final sheet found for: {sep_sheet}")

        # Check if all final sheets have matching separate files  
        missing_separate_files = []
        for canon, final_sheet in final_canon_map.items():
            sep_sheet = separate_canon_map.get(canon)
            if sep_sheet is None:
                missing_separate_files.append(final_sheet)
                print(f"❌ No separate file found for final sheet: {final_sheet}")

        # If there are any mismatches, return error immediately
        if missing_final_sheets or missing_separate_files:
            print(f"\n🚨 SHEET MATCHING ERROR:")
            print(f"   Final file has {len(final_canon_map)} sheets")
            print(f"   You uploaded {len(separate_canon_map)} files")
            
            if missing_final_sheets:
                print(f"   ❌ Missing final sheets for {len(missing_final_sheets)} uploaded files:")
                for file in missing_final_sheets:
                    print(f"      - {file}")
            
            if missing_separate_files:
                print(f"   ❌ Missing uploaded files for {len(missing_separate_files)} final sheets:")
                for sheet in missing_separate_files:
                    print(f"      - {sheet}")
            
            # Return error response
            return Response(
                {
                    "error": f"Sheet matching failed. Final file has {len(final_canon_map)} sheets but you uploaded {len(separate_canon_map)} files.",
                    "details": {
                        "final_sheets_count": len(final_canon_map),
                        "uploaded_files_count": len(separate_canon_map),
                        "missing_final_sheets_count": len(missing_final_sheets),
                        "missing_separate_files_count": len(missing_separate_files),
                        "missing_final_sheets_for_files": missing_final_sheets,
                        "missing_files_for_final_sheets": missing_separate_files,
                        "all_final_sheets": list(final_canon_map.values()),
                        "all_uploaded_files": list(separate_canon_map.values())
                    }
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        # If we reach here, all sheets are matched
        print(f"✅ ALL SHEETS MATCHED SUCCESSFULLY!")
        print(f"   Final sheets: {len(final_canon_map)}")
        print(f"   Uploaded files: {len(separate_canon_map)}")
        print(f"   Starting comparison process...\n")

        # NOW START THE COMPARISON
        print("=== STARTING PROCESSED vs FINAL COMPARISON ===")
        print(f"Total comparisons to make: {len(separate_canon_map)}")

        comparisons_made = []
        any_mismatch = False

        for canon, sep_sheet in separate_canon_map.items():
            final_sheet = final_canon_map.get(canon)
            
            # Since we already validated, these should never be None
            auto_df_full = all_files_to_process.get(sep_sheet)
            manual_df_full = df_storage.get(final_sheet)

            # Add to comparisons made list
            comparisons_made.append(sep_sheet)
            
            print(f"\n🔍 COMPARISON {len(comparisons_made)}/{len(separate_canon_map)}: {sep_sheet} vs {final_sheet}")
            
            # Compare INB TRF/SIS CON rows (using filtered data for comparison)
            comparison_result = compare_inb_sis_rows(auto_df_full, manual_df_full)

            sheet_id = f"{final_file_label}::{final_sheet}"
            mismatch_summary[sheet_id] = {
                "INB TRF/SIS CON Comparison": comparison_result
            }

            print(f"  Processed INB TRF rows: {comparison_result['auto_inb_count']}")
            print(f"  Final INB TRF/SIS CON rows: {comparison_result['manual_inb_count']}")
            print(f"  Rows only in processed (green): {comparison_result['auto_only_rows']}")
            print(f"  Rows only in final (red): {comparison_result['manual_only_rows']}")

            if comparison_result['auto_only_count'] > 0 or comparison_result['manual_only_count'] > 0:
                any_mismatch = True

            # NOW WORK WITH FULL PROCESSED FILES (NO FILTER) FOR HIGHLIGHTING
            
            # 1. Highlight rows that are ONLY in processed file (GREEN)
            # These are actual row numbers in the full processed file
            auto_df_positions = [row_num - 2 for row_num in comparison_result['auto_only_rows'] 
                            if 0 <= row_num - 2 < len(auto_df_full)]
            
            if auto_df_positions:
                highlight_green_positions[sep_sheet] = set(auto_df_positions)
                print(f"  🟢 Highlighting {len(auto_df_positions)} rows in processed file (only in processed)")
            
            # 2. For rows only in final file, find the corresponding transactions in processed file
            # by looking at the full processed file and finding similar transactions
            if comparison_result['manual_only_count'] > 0:
                red_positions = []
                
                # Get the final-only rows data
                final_only_data = []
                for final_row_num in comparison_result['manual_only_rows']:
                    final_idx = final_row_num - 2
                    if 0 <= final_idx < len(manual_df_full):
                        final_row = manual_df_full.iloc[final_idx]
                        final_only_data.append({
                            'row_num': final_row_num,
                            'date': normalize_date(final_row["Date"]),
                            'desc': str(final_row["Description"]).strip().lower(),
                            'amount': final_row.get("DR", 0) or final_row.get("CR", 0)
                        })
                
                # Now search through FULL processed file (not just INB TRF) for similar transactions
                for final_data in final_only_data:
                    best_match_pos = None
                    best_match_score = 0
                    
                    for idx, processed_row in auto_df_full.iterrows():
                        processed_date = normalize_date(processed_row["Date"])
                        processed_desc = str(processed_row["Description"]).strip().lower()
                        processed_amount = processed_row.get("DR", 0) or processed_row.get("CR", 0)
                        
                        # Calculate match score
                        match_score = 0
                        
                        # Date match (highest priority)
                        if processed_date == final_data['date']:
                            match_score += 3
                        
                        # Amount match
                        if abs(float(processed_amount or 0) - abs(float(final_data['amount'] or 0))) <= AMOUNT_TOLERANCE:
                            match_score += 2
                        
                        # Description similarity
                        if final_data['desc'] in processed_desc or processed_desc in final_data['desc']:
                            match_score += 1
                        
                        # Check for common keywords
                        common_words = set(final_data['desc'].split()) & set(processed_desc.split())
                        if len(common_words) >= 2:
                            match_score += 1
                        
                        if match_score > best_match_score:
                            best_match_score = match_score
                            best_match_pos = auto_df_full.index.get_loc(idx)
                    
                    # If we found a good match, highlight it in red
                    if best_match_pos is not None and best_match_score >= 2:
                        red_positions.append(best_match_pos)
                        print(f"  🔴 Found match for final row {final_data['row_num']} at processed position {best_match_pos} (score: {best_match_score})")
                
                if red_positions:
                    highlight_red_positions[sep_sheet] = set(red_positions)
                    print(f"  🔴 Highlighting {len(red_positions)} rows in processed file (potential matches for final-only rows)")

        print(f"\n✅ COMPARISON COMPLETED SUCCESSFULLY!")
        print(f"   Total comparisons made: {len(comparisons_made)}")
        print(f"   Files compared: {comparisons_made}")
        print(f"   Data mismatches found: {'Yes' if any_mismatch else 'No'}")

        # now save processed files with mismatch highlight (only in processed files)
        save_matched_with_styles(
            all_files_to_process, acc_name_storage, highlight_red_positions, highlight_green_positions
        )

        # Generate the summary report
        summary_report = generate_summary_report(all_files_to_process, df_storage, separate_canon_map, final_canon_map, final_file_label, acc_name_storage)

        # DEBUG: Check summary report before Google Sheets
        print("=== BEFORE GOOGLE SHEETS UPDATE ===")
        print(f"Summary report has {len(summary_report)} files:")
        for item in summary_report:
            print(f"   - {item['File Name']}")

        # Update Google Sheets with the report data
        sheets_update_success = update_google_sheets(summary_report, final_file_label)

        # Update the final response
        response_data = {
            "success": True,
            "message": (
                "Processed files saved in Downloads/Matched_Statemants. "
                + (
                    "Mismatches detected in INB TRF/SIS CON rows."
                    if any_mismatch
                    else "No INB TRF/SIS CON mismatches between processed and final."
                )
                + f" Google Sheets update: {'Successful' if sheets_update_success else 'Failed'}"
            ),
            "processed_dir": str(get_downloads() / PROCESSED_DIR),
            "has_mismatch": any_mismatch,
            "row_count_summary": row_count_summary,
            "mismatch_summary": mismatch_summary,
            "google_sheets_updated": sheets_update_success,
            "files_processed": len(all_files_to_process),
            "files_in_summary": len(summary_report)
        }

        return Response(response_data, status=status.HTTP_200_OK)
